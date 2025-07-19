from flask import request, Blueprint, render_template, current_app, flash, redirect, url_for, jsonify, g
from datetime import datetime, timedelta
from flask_login import login_required
from sqlalchemy.exc import SQLAlchemyError
from app import db, csrf, shop_access_required
from app.models import Product, StockLog, Category, Sale
from app.utils.calculations.product_calculations import *
from app.utils.calculations.report_calculations import *
from app.utils.calculations.report_calculations import MonthlySalesAnalyzer

import logging

reports_bp = Blueprint('reports', __name__)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

@reports_bp.route('/shops/<int:shop_id>/products/<int:product_id>/report')
@login_required
@shop_access_required
def product_report(shop_id, product_id):
    product = Product.query.filter_by(id=product_id, shop_id=shop_id).first_or_404()
    time_period = request.args.get('time_period', 'month')
    page = request.args.get('page', 1, type=int)  

    analytics = {
        'total_revenue': calculate_total_revenue(product_id, time_period),
        'revenue_trend': calculate_revenue_trend(product_id, time_period),
        'total_units_sold': calculate_total_units_sold(product_id, time_period),
        'sales_trend': calculate_sales_trend(product_id, time_period),
        'avg_profit_margin': calculate_avg_profit_margin(product_id, time_period),
        'margin_trend': calculate_margin_trend(product_id, time_period),
        'peak_sales_day': get_peak_sales_day(product_id, time_period),
        'avg_days_between_sales': get_avg_days_between_sales(product_id, time_period),
        'max_stock_observed': get_max_stock_observed(product_id),
        'stockout_count': get_stockout_count(product_id, time_period),
        'avg_monthly_usage': get_avg_monthly_usage(product_id),
        'stock_cover_days': get_stock_cover_days(product_id),
        'best_selling_month': get_best_selling_month(product_id),
        'revenue_growth': get_revenue_growth(product_id, time_period),
        'sales_growth': get_sales_growth(product_id, time_period),
        'price_change_count': get_price_change_count(product_id, time_period),
        'suggested_price': get_suggested_price(product_id),
        'avg_quantity_per_order': get_avg_quantity_per_order(product_id, time_period),
        'repeat_purchase_rate': get_repeat_purchase_rate(product_id, time_period),
        'frequently_bought_with': get_frequently_bought_with(product_id, time_period),
        'months': get_analytics_months(product_id),
        'units_sold_by_month': get_units_sold_by_month(product_id),
        'revenue_by_month': get_revenue_by_month(product_id),
        'price_change_dates': get_price_change_dates(product_id),
        'price_history': get_price_history(product_id),
        'sales_by_day_of_week': get_sales_by_day_of_week(product_id, time_period)
    }

    return render_template(
        'reports/fragments/_product_analytics_dashboard.html',
        product=product,
        product_id=product_id,
        shop_id=shop_id,
        page=page,  
        time_period=time_period,
        analytics=analytics
    )




@reports_bp.route('/shops/<int:shop_id>/products/<int:product_id>/stock_history')
@login_required
@shop_access_required
def product_stock_history(shop_id, product_id):
    try:
        product = Product.query.filter_by(id=product_id, shop_id=shop_id).first_or_404()
        page = request.args.get('page', 1, type=int)
        per_page = 10

        stock_query = StockLog.query.filter_by(product_id=product.id).order_by(StockLog.date.desc())
        paginated_logs = stock_query.paginate(page=page, per_page=per_page)

        if request.headers.get('HX-Request'):
            return render_template('reports/fragments/_product_stock_history.html',
                                   stock_history=paginated_logs.items,
                                   pagination=paginated_logs,
                                   product_id=product.id)

        return render_template('reports/product_stock_history_full.html',
                               stock_history=paginated_logs.items,
                               pagination=paginated_logs,
                               product_id=product.id)

    except Exception as e:
        logger.error(f"Error loading stock history: {str(e)}")
        return render_template('reports/product_stock_history_full.html',
                               stock_history=[],
                               error_message="Could not load stock history",
                               pagination=None,
                               product_id=product_id)






@reports_bp.route('/shops/<int:shop_id>/products/<int:product_id>/price_history')
@login_required
@shop_access_required
def product_price_history(shop_id, product_id):
    price_history = PriceChange.query.filter_by(
        product_id=product_id
    ).order_by(
        PriceChange.changed_at.desc()
    ).limit(50).all()
    
    return render_template('reports/fragments/_product_price_history.html',
                        price_history=price_history)




@reports_bp.route('/shops/<int:shop_id>/products/<int:product_id>/sales_table')
@login_required
@shop_access_required
def product_sales_table(shop_id, product_id):
    page = request.args.get('page', 1, type=int)
    per_page = 10
    time_period = request.args.get('time_period', 'month')

    try:
        product = Product.query.get(product_id)
        if not product:
            return render_template('reports/fragments/_product_sales_table.html',
                                   sales_records=[],
                                   product_id=product_id,
                                   page=page,
                                   total_pages=1,
                                   time_period=time_period,
                                   total_sales=0,
                                   error_message="Product not found")

        now = datetime.utcnow()
        if time_period == 'day':
            start_date = now - timedelta(days=1)
        elif time_period == 'week':
            start_date = now - timedelta(weeks=1)
        elif time_period == 'month':
            start_date = now - timedelta(days=30)
        elif time_period == 'year':
            start_date = now - timedelta(days=365)
        else:
            start_date = now - timedelta(days=30)

        sales_query = db.session.query(
            Sale,
            CartItem.quantity,
            (CartItem.quantity * Product.selling_price).label('total_price'),
            Product.selling_price
        ).join(
            CartItem, and_(
                Sale.id == CartItem.sale_id,
                CartItem.product_id == product_id
            )
        ).join(
            Product, CartItem.product_id == Product.id
        ).filter(
            Sale.date >= start_date
        ).order_by(
            Sale.date.desc()
        )

        paginated_sales = sales_query.paginate(page=page, per_page=per_page, error_out=False)
        total_records = paginated_sales.total

        sales_data = []
        for sale, quantity, total_price, unit_price in paginated_sales.items:
            sales_data.append({
                'sale_date': sale.date.strftime('%Y-%m-%d %I:%M %p') if sale.date else '',
                'invoice_id': sale.id,
                'invoice_number': f"INV-{sale.id:05d}",
                'quantity': quantity,
                'unit_price': float(product.selling_price) if product and product.selling_price else 0.0,
                'discount': 0.00,
                'total_price': float(total_price),
                'product_name': product.name if product else "N/A",
                'payment_method': sale.payment_method or "-",        
                'customer_name': sale.customer_name or "-"
            })




        return render_template('reports/fragments/_product_sales_table.html',
                                sales_records=sales_data,
                                product_id=product_id,
                                shop_id=shop_id,  # ✅ include this
                                page=page,
                                total_pages=paginated_sales.pages,
                                time_period=time_period,
                                total_sales=total_records)


    except Exception as e:
        logger.exception("Error loading sales data")
        return render_template('reports/fragments/_product_sales_table.html',
                                sales_records=[],
                                product_id=product_id,
                                shop_id=shop_id,  # ✅ include this
                                page=page,
                                total_pages=1,
                                time_period=time_period,
                                total_sales=0,
                                error_message="Could not load sales data")

@reports_bp.route('/shops/<int:shop_id>/api/todays-total-sales', methods=['GET'])
@login_required
@shop_access_required
def todays_total_sales(shop_id):
    """Fetch today's total sales and number of transactions for this shop."""
    today = datetime.today().date()

    try:
        total_sales, total_transactions = db.session.query(
            func.coalesce(func.sum(Sale.total), 0),
            func.count(Sale.id)
        ).filter(
            func.date(Sale.date) == today,
            Sale.shop_id == shop_id  
        ).first()

        return jsonify({
            'total_sales': round(total_sales, 2),
            'total_transactions': total_transactions
        })

    except SQLAlchemyError as e:
        logging.error(f"[Shop {shop_id}] Error fetching today's sales: {e}")
        return jsonify({'error': 'Failed to fetch sales data'}), 500



@reports_bp.route('/shops/<int:shop_id>/reports/daily', methods=['GET'])
@login_required
@shop_access_required
def daily_sales_report(shop_id):
    try:
        date_str = request.args.get('date', datetime.today().strftime('%Y-%m-%d'))
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        if report_date > datetime.today().date():
            raise ValueError("Future dates not allowed")
        if report_date < (datetime.today() - timedelta(days=730)).date():
            raise ValueError("Date too far in the past")

    except ValueError as e:
        if request.headers.get('HX-Request'):
            return render_template('sales/fragments/_error_message.html', message=str(e)), 400
        if request.accept_mimetypes.accept_json:
            return jsonify({"error": str(e)}), 400
        flash(str(e), "danger")
        return redirect(url_for('reports.daily_sales_report', shop_id=shop_id))

    cache_key = f"shop:{shop_id}:daily_report:{report_date}"
    if request.headers.get("HX-Request"):
        cached = cache.get(cache_key)
        if cached:
            return cached

    report_data = generate_daily_report_data(shop_id=shop_id, report_date=report_date)

    # JSON response
    if request.accept_mimetypes.best == 'application/json':
        serialized = {
            **report_data,
            'sales': [
                {
                    "id": sale.id,
                    "user": sale.user.username if sale.user else None,
                    "total": float(sale.total),
                    "profit": float(sale.profit or 0),
                    "payment_method": sale.payment_method,
                    "date": sale.date.strftime('%Y-%m-%d %H:%M:%S'),
                    "items": [
                        {
                            "product": item.product.name,
                            "quantity": item.quantity,
                            "total_price": float(item.total_price)
                        }
                        for item in sale.cart_items
                    ]
                }
                for sale in report_data.get("sales", [])
            ]
        }
        return jsonify(serialized)

    return render_htmx(
        "reports/fragments/_daily_report.html",
        shop_id=shop_id,
        timedelta=timedelta,
        datetime=datetime,
        **report_data
    )


@reports_bp.route('/shops/<int:shop_id>/reports/weekly', methods=['GET'])
@login_required
@shop_access_required
def weekly_sales_report(shop_id):
    try:
        # Get week string, fallback to current week
        week = request.args.get('week', datetime.today().strftime('%Y-W%W'))
        year, week_num = map(int, week.split('-W'))
        start_date = datetime.strptime(f'{year}-W{week_num}-1', "%Y-W%W-%w").date()
        end_date = start_date + timedelta(days=6)
    except ValueError:
        # Fallback if parsing fails
        today = datetime.today().date()
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        week = start_date.strftime('%Y-W%W')

    # 🚨 Make sure this function is shop-scoped!
    context = generate_weekly_report_context(shop_id=shop_id, week=week, start_date=start_date, end_date=end_date)


    # HTMX request (partial fragment)
    if request.headers.get("HX-Request"):
        return render_template("reports/fragments/_weekly_sales_fragment.html", **context)

    # Full render wrapper
    return render_htmx(
        "reports/fragments/_weekly_sales_fragment.html",
        shop_id=shop_id,
        **context
    )

        
@reports_bp.route('/shops/<int:shop_id>/reports/monthly-analytics', methods=['GET'])
@login_required
@shop_access_required
def monthly_sales_analytics(shop_id):
    """Monthly Sales Analytics scoped by shop"""
    try:
        month = request.args.get('month')  # format: 'YYYY-MM'
        if not month:
            month = datetime.today().strftime('%Y-%m')  # default to current month

        analyzer = MonthlySalesAnalyzer(shop_id=shop_id, month_str=month)

        context = analyzer.generate_context(full_report=not request.headers.get('HX-Request'))

        return render_htmx(
            'reports/fragments/_monthly_sales_fragment.html',
            **context
        )

    except Exception as e:
        current_app.logger.error(f"[Monthly Analytics] shop_id={shop_id} error: {str(e)}", exc_info=True)

        if request.headers.get('HX-Request'):
            return render_template(
                'admin/fragments/_error.html',
                message=f"Failed to load monthly analytics: {str(e)}"
            ), 500

        flash("An error occurred while generating the monthly report. Please try again.", "danger")
        return redirect(url_for('reports.daily_sales_report', shop_id=shop_id))



from flask import make_response, request, current_app
from flask_login import login_required
from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Register professional fonts (fallback to Helvetica if not available)
try:
    pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Regular.ttf'))
    pdfmetrics.registerFont(TTFont('Roboto-Bold', 'Roboto-Bold.ttf'))
    pdfmetrics.registerFont(TTFont('Roboto-Light', 'Roboto-Light.ttf'))
except:
    pass  # Fallback to default fonts

@reports_bp.route('/reports/daily/export-pdf', methods=['GET'])
@login_required
def export_daily_report_pdf():
    try:
        # Date handling
        date_str = request.args.get('date', datetime.today().strftime('%Y-%m-%d'))
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        formatted_date = report_date.strftime("%A, %B %d, %Y").upper()
        
        # Data generation
        report_data = generate_daily_report_data(report_date)
        summary = report_data.get('summary', {})
        
        # PDF setup with tighter margins
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=15*mm,
            rightMargin=15*mm,
            topMargin=10*mm,
            bottomMargin=15*mm,
            title=f"Daily Sales Report - {formatted_date}"
        )
        elements = []

        # Custom Styles
        styles = getSampleStyleSheet()
        
        # Title style
        styles.add(ParagraphStyle(
            name='ReportTitle',
            fontName='Helvetica-Bold',
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=6,
            textColor=colors.HexColor("#2c3e50")
        ))
        
        # Date style
        styles.add(ParagraphStyle(
            name='ReportDate',
            fontName='Helvetica',
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=18,
            textColor=colors.HexColor("#7f8c8d")
        ))
        
        # Section header style
        styles.add(ParagraphStyle(
            name='SectionHeader',
            fontName='Helvetica-Bold',
            fontSize=12,
            textColor=colors.HexColor("#3498db"),
            spaceAfter=8,
            underlineWidth=1,
            underlineColor=colors.HexColor("#3498db"),
            underlineOffset=-3
        ))
        
        # Table header style
        styles.add(ParagraphStyle(
            name='TableHeader',
            fontName='Helvetica-Bold',
            fontSize=9,
            alignment=TA_CENTER,
            textColor=colors.white
        ))
        
        # Body text style
        styles.add(ParagraphStyle(
            name='RBodyText',
            fontName='Helvetica',
            fontSize=9,
            leading=11,
            spaceAfter=6
        ))
        
        # Footer style
        styles.add(ParagraphStyle(
            name='FooterText',
            fontName='Helvetica-Oblique',
            fontSize=8,
            textColor=colors.HexColor("#95a5a6"),
            alignment=TA_CENTER
        ))

        # Report Header
        elements.append(Paragraph("DAILY SALES REPORT", styles['ReportTitle']))
        elements.append(Paragraph(formatted_date, styles['ReportDate']))
        elements.append(HRFlowable(width="80%", thickness=0.5, lineCap='round', 
                                 color=colors.HexColor("#bdc3c7"), spaceAfter=18))

        # SECTION 1: Summary (Card-style layout)
        elements.append(Paragraph("PERFORMANCE SUMMARY", styles['SectionHeader']))
        
        summary_data = [
            ("TOTAL SALES", f"Ksh {summary.get('total_sales', 0):,.2f}", "#2ecc71"),
            ("TOTAL TRANSACTIONS", f"{summary.get('total_transactions', 0):,}", "#3498db"),
            ("AVG SALE", f"Ksh {summary.get('avg_sale', 0):,.2f}", "#9b59b6"),
            ("TOTAL PROFIT", f"Ksh {summary.get('total_profit', 0):,.2f}", "#e74c3c")
        ]
        
        summary_cards = []
        for title, value, color in summary_data:
            card = Table([
                [Paragraph(title, ParagraphStyle(
                    name='CardTitle',
                    fontName='Helvetica-Bold',
                    fontSize=9,
                    textColor=colors.white,
                    alignment=TA_CENTER
                ))],
                [Paragraph(value, ParagraphStyle(
                    name='CardValue',
                    fontName='Helvetica-Bold',
                    fontSize=11,
                    textColor=colors.white,
                    alignment=TA_CENTER
                ))]
            ], colWidths=[2.25*inch], rowHeights=[0.3*inch, 0.4*inch])
            
            card.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(color)),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.white),
                ('ROUNDEDCORNERS', [4, 4, 4, 4]),
            ]))
            summary_cards.append(card)
        
        # Arrange cards in 2x2 grid
        summary_grid = Table([
            [summary_cards[0], summary_cards[1]],
            [summary_cards[2], summary_cards[3]]
        ], colWidths=[2.5*inch, 2.5*inch], rowHeights=[0.8*inch, 0.8*inch])
        
        elements.append(KeepTogether(summary_grid))
        elements.append(Spacer(1, 0.3*inch))

        # SECTION 2: Payment Methods
        payments = report_data.get('payment_methods', {})
        elements.append(Paragraph("PAYMENT METHODS", styles['SectionHeader']))
        
        if payments:
            payment_data = [["METHOD", "AMOUNT (Ksh)"]]
            payment_data.extend([
                [method.upper(), Paragraph(f"{amount:,.2f}", ParagraphStyle(
                    name='RightAlign',
                    fontName='Helvetica',
                    fontSize=9,
                    alignment=TA_RIGHT
                ))] 
                for method, amount in payments.items()
            ])
            
            payment_table = Table(
                payment_data,
                colWidths=[4*inch, 2*inch],
                repeatRows=1
            )
            
            payment_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#ecf0f1")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9f9")]),
            ]))
            
            elements.append(payment_table)
        else:
            elements.append(Paragraph("No payment data available.", styles['RBodyText']))
        
        elements.append(Spacer(1, 0.3*inch))

        # SECTION 3: Top Selling Products
        top_products = report_data.get('product_performance', [])
        elements.append(Paragraph("TOP SELLING PRODUCTS", styles['SectionHeader']))
        
        if top_products:
            product_data = [["PRODUCT", "QTY", "REVENUE (Ksh)"]]
            product_data.extend([
                [product, 
                 str(data['quantity']), 
                 Paragraph(f"{data['revenue']:,.2f}", ParagraphStyle(
                     name='RightAlign',
                     fontName='Helvetica',
                     fontSize=9,
                     alignment=TA_RIGHT
                 ))] 
                for product, data in top_products
            ])
            
            product_table = Table(
                product_data,
                colWidths=[3.5*inch, 1.25*inch, 1.25*inch],
                repeatRows=1
            )
            
            product_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#ecf0f1")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9f9")]),
            ]))
            
            elements.append(product_table)
        else:
            elements.append(Paragraph("No product sales recorded.", styles['RBodyText']))
        
        elements.append(Spacer(1, 0.3*inch))

        # SECTION 4: Low Stock Products
        from app.models import Product  # Import as needed, adjust path
        
        elements.append(Paragraph("LOW STOCK ALERTS", styles['SectionHeader']))
        low_stock_threshold = 10
        critical_stock_threshold = 5

        low_stock_products = Product.query.filter(
            Product.stock <= low_stock_threshold
        ).order_by(Product.stock.asc()).limit(10).all()

        if low_stock_products:
            low_stock_data = [["PRODUCT", "STOCK", "REORDER LEVEL", "CATEGORY"]]
            for p in low_stock_products:
                stock_style = 'Helvetica-Bold' if p.stock <= critical_stock_threshold else 'Helvetica'
                stock_color = colors.red if p.stock <= critical_stock_threshold else colors.black
                
                low_stock_data.append([
                    p.name,
                    Paragraph(str(p.stock), ParagraphStyle(
                        name='StockAlert',
                        fontName=stock_style,
                        fontSize=9,
                        textColor=stock_color,
                        alignment=TA_CENTER
                    )),
                    str(getattr(p, 'reorder_level', 10)),
                    p.category.name if p.category else "N/A"
                ])

            stock_table = Table(
                low_stock_data,
                colWidths=[2.5*inch, 0.75*inch, 1*inch, 1.25*inch],
                repeatRows=1
            )
            
            stock_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#ecf0f1")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9f9")]),
            ]))
            
            elements.append(stock_table)
        else:
            elements.append(Paragraph("All products are sufficiently stocked.", styles['RBodyText']))
        
        elements.append(Spacer(1, 0.3*inch))

        # SECTION 5: Staff Performance
        staff_performance = report_data.get('staff_performance', [])
        elements.append(Paragraph("STAFF PERFORMANCE", styles['SectionHeader']))
        
        if staff_performance:
            staff_data = [["STAFF MEMBER", "TRANSACTIONS", "SALES (Ksh)"]]
            staff_data.extend([
                [staff, 
                 str(data['sales']), 
                 Paragraph(f"{data['amount']:,.2f}", ParagraphStyle(
                     name='RightAlign',
                     fontName='Helvetica',
                     fontSize=9,
                     alignment=TA_RIGHT
                 ))] 
                for staff, data in staff_performance
            ])
            
            staff_table = Table(
                staff_data,
                colWidths=[3*inch, 1.5*inch, 1.5*inch],
                repeatRows=1
            )
            
            staff_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#ecf0f1")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9f9")]),
            ]))
            
            elements.append(staff_table)
        else:
            elements.append(Paragraph("No staff performance data available.", styles['RBodyText']))
        
        elements.append(Spacer(1, 0.3*inch))

        # SECTION 6: Hourly Trends
        hourly_trends = report_data.get('hourly_trends', [])
        elements.append(Paragraph("HOURLY SALES TRENDS", styles['SectionHeader']))
        
        if hourly_trends:
            hourly_data = [["HOUR", "SALES (Ksh)"]]
            hourly_data.extend([
                [hour, 
                 Paragraph(f"{amount:,.2f}", ParagraphStyle(
                     name='RightAlign',
                     fontName='Helvetica',
                     fontSize=9,
                     alignment=TA_RIGHT
                 ))] 
                for hour, amount in hourly_trends
            ])
            
            hourly_table = Table(
                hourly_data,
                colWidths=[3*inch, 3*inch],
                repeatRows=1
            )
            
            hourly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#ecf0f1")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9f9")]),
            ]))
            
            elements.append(hourly_table)
        else:
            elements.append(Paragraph("No hourly sales data available.", styles['RBodyText']))
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))
        elements.append(HRFlowable(width="100%", thickness=0.5, lineCap='round', 
                                 color=colors.HexColor("#bdc3c7"), spaceAfter=6))
        elements.append(Paragraph(
            f"Generated on {datetime.now().strftime('%Y-%m-%d at %H:%M:%S')} • © {datetime.now().year} Nawiri Enterprise",
            styles['FooterText']
        ))

        # Build and return PDF
        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = (
            f'attachment; filename=daily_sales_report_{report_date}.pdf'
        )
        return response

    except Exception as e:
        current_app.logger.error(f"PDF generation error: {str(e)}")
        return str(e), 500




from flask import make_response, request, current_app
from flask_login import login_required
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import os

@reports_bp.route('/reports/daily/export-excel', methods=['GET'])
@login_required
def export_daily_report_excel():
    try:
        date_str = request.args.get('date', datetime.today().strftime('%Y-%m-%d'))
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        formatted_date = report_date.strftime("%B %d, %Y")

        report_data = generate_daily_report_data(report_date)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"
        
        # Set default column width
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Create styles
        header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(name='Calibri', bold=True, size=14)
        title_style.alignment = Alignment(horizontal='center')
        
        section_style = NamedStyle(name="section_style")
        section_style.font = Font(name='Calibri', bold=True, size=12, color='3498DB')
        section_style.alignment = Alignment(horizontal='left')
        
        currency_style = NamedStyle(name="currency_style")
        currency_style.number_format = '"Ksh" #,##0.00'
        currency_style.alignment = Alignment(horizontal='right')
        
        # Add styles to workbook
        wb.add_named_style(title_style)
        wb.add_named_style(section_style)
        wb.add_named_style(currency_style)

       

        # Report title
        ws['A3'] = "DAILY SALES REPORT"
        ws['A3'].style = title_style
        ws.merge_cells('A3:D3')
        
        ws['A4'] = formatted_date
        ws['A4'].font = Font(name='Calibri', italic=True)
        ws.merge_cells('A4:D4')
        
        current_row = 6

        # SECTION 1: Summary
        summary = report_data.get('summary', {})
        ws.cell(row=current_row, column=1, value="PERFORMANCE SUMMARY").style = section_style
        current_row += 1
        
        summary_headers = ["Metric", "Value"]
        ws.append(summary_headers)
        
        for cell in ws[current_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        summary_data = [
            ["Total Sales", summary.get('total_sales', 0)],
            ["Total Transactions", summary.get('total_transactions', 0)],
            ["Average Sale", summary.get('avg_sale', 0)],
            ["Total Profit", summary.get('total_profit', 0)]
        ]
        
        for row in summary_data:
            ws.append(row)
            ws.cell(row=current_row+1, column=2).style = currency_style
        
        # Apply borders to summary data
        for row in ws.iter_rows(min_row=current_row, max_row=current_row+3, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
        
        current_row += 5

        # SECTION 2: Payment Methods
        payments = report_data.get('payment_methods', {})
        ws.cell(row=current_row, column=1, value="PAYMENT METHODS").style = section_style
        current_row += 1
        
        payment_headers = ["Method", "Amount (Ksh)"]
        ws.append(payment_headers)
        
        for cell in ws[current_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for method, amount in payments.items():
            ws.append([method.capitalize(), amount])
            ws.cell(row=current_row+1, column=2).style = currency_style
            current_row += 1
        
        # Apply borders to payment data
        for row in ws.iter_rows(min_row=current_row-len(payments), max_row=current_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
        
        current_row += 2

        # SECTION 3: Top Selling Products
        top_products = report_data.get('product_performance', [])
        ws.cell(row=current_row, column=1, value="TOP SELLING PRODUCTS").style = section_style
        current_row += 1
        
        product_headers = ["Product", "Quantity", "Revenue (Ksh)"]
        ws.append(product_headers)
        
        for cell in ws[current_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for product, data in top_products:
            ws.append([product, data['quantity'], data['revenue']])
            ws.cell(row=current_row+1, column=3).style = currency_style
            current_row += 1
        
        # Apply borders to product data
        for row in ws.iter_rows(min_row=current_row-len(top_products), max_row=current_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
        
        current_row += 2

        # SECTION 4: Low Stock Products
        from app.models import Product
        ws.cell(row=current_row, column=1, value="LOW STOCK ALERTS").style = section_style
        current_row += 1
        
        low_stock_headers = ["Product", "Stock", "Reorder Level", "Category"]
        ws.append(low_stock_headers)
        
        for cell in ws[current_row]:
            cell.font = header_font
            cell.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = thin_border
        
        low_stock_threshold = 10
        critical_stock_threshold = 5
        low_stock_products = Product.query.filter(
            Product.stock <= low_stock_threshold
        ).order_by(Product.stock.asc()).limit(10).all()

        for p in low_stock_products:
            ws.append([
                p.name,
                p.stock,
                getattr(p, 'reorder_level', 10),
                p.category.name if p.category else "N/A"
            ])
            # Highlight critical stock in red
            if p.stock <= critical_stock_threshold:
                ws.cell(row=current_row+1, column=2).font = Font(color='E74C3C', bold=True)
            current_row += 1
        
        # Apply borders to stock data
        for row in ws.iter_rows(min_row=current_row-len(low_stock_products), max_row=current_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
        
        current_row += 2

        # SECTION 5: Staff Performance
        staff_performance = report_data.get('staff_performance', [])
        ws.cell(row=current_row, column=1, value="STAFF PERFORMANCE").style = section_style
        current_row += 1
        
        staff_headers = ["Staff Member", "Transactions", "Sales (Ksh)"]
        ws.append(staff_headers)
        
        for cell in ws[current_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for staff, data in staff_performance:
            ws.append([staff, data['sales'], data['amount']])
            ws.cell(row=current_row+1, column=3).style = currency_style
            current_row += 1
        
        # Apply borders to staff data
        for row in ws.iter_rows(min_row=current_row-len(staff_performance), max_row=current_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
        
        current_row += 2

        # SECTION 6: Hourly Trends
        hourly_trends = report_data.get('hourly_trends', [])
        ws.cell(row=current_row, column=1, value="HOURLY SALES TRENDS").style = section_style
        current_row += 1
        
        hourly_headers = ["Hour", "Sales (Ksh)"]
        ws.append(hourly_headers)
        
        for cell in ws[current_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for hour, amount in hourly_trends:
            ws.append([hour, amount])
            ws.cell(row=current_row+1, column=2).style = currency_style
            current_row += 1
        
        # Apply borders to hourly data
        for row in ws.iter_rows(min_row=current_row-len(hourly_trends), max_row=current_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
        
        # Footer
        current_row += 2
        ws.cell(row=current_row, column=1, 
                value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color='7F8C8D')
        
        ws.cell(row=current_row, column=4, value="Confidential")
        ws.cell(row=current_row, column=4).font = Font(italic=True, color='7F8C8D')
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='right')

        # Freeze headers
        ws.freeze_panes = 'A7'

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = make_response(buffer.read())
        response.headers['Content-Type'] = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = (
            f'attachment; filename=daily_sales_report_{report_date}.xlsx'
        )
        return response

    except Exception as e:
        current_app.logger.error(f"Excel generation error: {str(e)}")
        return str(e), 500